import openpyxl
import csv
import os

# 获取当前目录下的所有文件，找到后缀为.csv的文件并获取文件名
list_pwd = os.listdir(os.getcwd())
for csv_file in list_pwd:
    if '.csv' in csv_file:
        break
filenpathame_csv = csv_file
filenpathame_xlsx = 'metersphere导入文件.xlsx'
sheetname = 'Sheet'
# 标记用例状态
case_status = "Completed"


def csv2xlsx():
    """
    csv转为xlsx excel2007以上版本
    """
    with open(filenpathame_csv, 'r', encoding='utf-8') as f:
        read = csv.reader(f)
        wb = openpyxl.Workbook()
        ws = wb.active
        # 将读取的csv文件原格式保存到xlsx文件
        for line in read:
            ws.append(line)
        wb.save(filenpathame_xlsx)  # 保存Excel


def fill_xlsx():
    """
    修改xlsx列字段，并将coding的导出用例如操作步骤和预计结果与metersphere进行格式对齐
    """
    wb = openpyxl.load_workbook(filenpathame_xlsx)
    ws = wb[sheetname]
    rows = ws.max_row
    columns = ws.max_column
    # 修改列名，将无用列挑出来
    ws["A1"] = "所属模块"
    ws["B1"] = "废弃2"
    ws["C1"] = "用例名称"
    ws["D1"] = "废弃"
    ws["E1"] = "废弃"
    ws["F1"] = "用例等级"
    ws["G1"] = "前置条件"
    ws["H1"] = "废弃"
    ws["I1"] = "废弃"
    ws["J1"] = "步骤描述"
    ws["K1"] = "预期结果"
    ws["L1"] = "废弃"
    ws["M1"] = "用例状态"
    ws["N1"] = "废弃14"
    ws["O1"] = "废弃15"
    ws["P1"] = "用例状态"
    # 删除无用列
    ws.delete_cols(2)
    ws.delete_cols(3, 2)
    ws.delete_cols(5, 2)
    ws.delete_cols(7, 4)
    ws.delete_cols(8)
    # 修正列表数据，如 状态，操作步骤格式等
    for i in range(2, rows + 1):
        ws.cell(row=i, column=7, value=case_status)
        ws.cell(row=i, column=5, value=format_numbers_with_brackets(remove_new_lines(ws.cell(row=i, column=5).value)))
        ws.cell(row=i, column=6, value=format_numbers_with_brackets(remove_new_lines(ws.cell(row=i, column=6).value)))
    print("转换完成")
    wb.save(filenpathame_xlsx)


def format_numbers_with_brackets(string):
    """
    用于更改步骤和预期结果的格式
    :param string: 文本内容，如操作步骤，预期结果等
    :return: 修正后的文本内容
    """
    result = ""
    for i in range(len(string)):
        char = string[i]
        # 判断是否为数字，如果是则下一步判断
        if char.isdigit():
            # 判断前一位是不是[和后一位是不是]如果同时满足进行下一步判断，如果不是则将该字符存入result字符串
            if 0 < i < len(string) - 1 and string[i - 1] == "[" and string[i + 1] == "]":
                # 判断是不是首位，如果是则替换成空，如果不是则替换成"\n"
                if i - 1 != 0:
                    result += f"\n"
                else:
                    result += ""
                i += 1
            else:
                result += char
        # 判断如果是[数字]格式，则[和]不保留
        elif char == "[" and string[i + 1].isdigit() or char == "]" and string[i - 1].isdigit():
            continue
        else:
            result += char
    return result


def remove_new_lines(string):
    """
    将字符串中的回车全部去掉（针对部分用例单条步骤中有回车对导入的影响）
    :param string:字符串类型，如操作步骤，预期结果等
    :return:返回去除后的文本内容
    """
    return string.replace("\n", "")


if __name__ == '__main__':
    csv2xlsx()
    fill_xlsx()

# Press the green button in the gutter to run the script.
